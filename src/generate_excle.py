import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

def create_sales_rep_reports():
    print("Sales Rep Excel Report Generator")
    print("=" * 50)
    conn = duckdb.connect(':memory:')
    try:
        print("Loading CSV data...")
        conn.execute("""
            CREATE TABLE sales_data AS 
            SELECT * FROM read_csv_auto('resources/assignmentData.csv')
        """)
        overview = conn.execute("""
            SELECT 
                COUNT(*) as total_records,
                COUNT(DISTINCT SALES_REP_ID) as unique_sales_reps,
                COUNT(DISTINCT CUSTOMER_ID) as unique_customers
            FROM sales_data
            WHERE TOTAL_VALUE_SOLD IS NOT NULL
        """).fetchone()
        print(f"Loaded {overview[0]:,} records")
        print(f"   Sales Reps: {overview[1]}, Customers: {overview[2]}")
        print("\nFinding Top 3 Sales Representatives...")
        top3_query = """
        SELECT 
            SALES_REP_ID,
            SALES_REP,
            SUM(TOTAL_VALUE_SOLD) as total_sales,
            COUNT(*) as total_transactions,
            ROW_NUMBER() OVER (ORDER BY SUM(TOTAL_VALUE_SOLD) DESC) as rank
        FROM sales_data
        WHERE TOTAL_VALUE_SOLD IS NOT NULL 
          AND SALES_REP IS NOT NULL
          AND TRIM(SALES_REP) != ''
        GROUP BY SALES_REP_ID, SALES_REP
        ORDER BY total_sales DESC
        LIMIT 3
        """
        top_reps = conn.execute(top3_query).fetchdf()
        print("\nTop 3 Sales Representatives:")
        for _, rep in top_reps.iterrows():
            print(f"  #{rep['rank']} {rep['SALES_REP']}: ${rep['total_sales']:,.2f}")
        created_files = []
        for _, rep in top_reps.iterrows():
            filename = generate_excel_report(conn, rep)
            created_files.append(filename)
        print(f"\nSuccessfully created {len(created_files)} Excel files!")
        print("\nFiles created:")
        for file in created_files:
            print(f"  {file}")
        return created_files
    except Exception as e:
        print(f"Error: {str(e)}")
        return []
    finally:
        conn.close()

def generate_excel_report(conn, rep_info):
    sales_rep_id = rep_info['SALES_REP_ID']
    sales_rep_name = rep_info['SALES_REP']
    rank = rep_info['rank']
    print(f"\nCreating report for Rep #{rank}: {sales_rep_name}")
    safe_name = sales_rep_name.replace(' ', '_').replace('/', '_')
    filename = f"SalesRep_{rank}_{safe_name}_Report.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    print("  Building Day-wise Transactions...")
    transactions_sql = f"""
    WITH daily_transactions AS (
        SELECT 
            CAST(strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M') AS DATE) as transaction_date,
            SKU_NAME,
            TOTAL_VALUE_SOLD / UNIT_SOLD as price_per_sku,
            UNIT_SOLD,
            TOTAL_VALUE_SOLD,
            ROW_NUMBER() OVER (
                PARTITION BY CAST(strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M') AS DATE)
                ORDER BY strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M')
            ) as sku_index
        FROM sales_data
        WHERE SALES_REP_ID = {sales_rep_id}
          AND SKU_NAME IS NOT NULL
          AND UNIT_SOLD > 0
          AND TOTAL_VALUE_SOLD > 0
          AND CHECKIN_TIME IS NOT NULL
    )
    SELECT 
        transaction_date as "Date",
        sku_index as "SKU_Index",
        SKU_NAME as "SKU_Sold", 
        ROUND(price_per_sku, 2) as "Price_per_SKU",
        UNIT_SOLD as "Quantity_Sold",
        ROUND(TOTAL_VALUE_SOLD, 2) as "Value_Sold"
    FROM daily_transactions
    ORDER BY transaction_date, sku_index
    """
    try:
        transactions_df = conn.execute(transactions_sql).fetchdf()
        create_transactions_sheet(wb, transactions_df)
        print(f"    Added {len(transactions_df)} transactions")
    except Exception as e:
        print(f"    Transaction sheet error: {e}")
        ws = wb.create_sheet("Day_wise_Transactions")
        ws['A1'] = "No transaction data available"
    print("  Building Summary Report...")
    summary_sql = f"""
    WITH daily_stats AS (
        SELECT 
            CAST(strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M') AS DATE) as report_date,
            SUM(UNIT_SOLD) as total_quantity,
            SUM(TOTAL_VALUE_SOLD) as total_value,
            COUNT(DISTINCT SKU_NAME) as unique_skus,
            COUNT(DISTINCT CUSTOMER_ID) as unique_customers,
            COUNT(*) as total_visits,
            SUM(
                CASE 
                    WHEN CHECKOUT_TIME IS NOT NULL AND CHECKIN_TIME IS NOT NULL
                    THEN EXTRACT(EPOCH FROM (
                        strptime(CHECKOUT_TIME, '%m/%d/%Y %H:%M') - 
                        strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M')
                    )) / 60.0
                    ELSE 0
                END
            ) as total_minutes
        FROM sales_data
        WHERE SALES_REP_ID = {sales_rep_id}
          AND CHECKIN_TIME IS NOT NULL
        GROUP BY CAST(strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M') AS DATE)
    ),
    conversion_data AS (
        SELECT 
            CAST(strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M') AS DATE) as conv_date,
            COUNT(DISTINCT CUSTOMER_ID) as total_customer_contacts,
            COUNT(DISTINCT CASE WHEN TOTAL_VALUE_SOLD > 0 THEN CUSTOMER_ID END) as customers_with_sales
        FROM sales_data
        WHERE SALES_REP_ID = {sales_rep_id}
          AND CHECKIN_TIME IS NOT NULL
        GROUP BY CAST(strptime(CHECKIN_TIME, '%m/%d/%Y %H:%M') AS DATE)
    )
    SELECT 
        ds.report_date as "Date",
        ds.total_quantity as "Total_Quantity_Sold",
        ROUND(ds.total_value, 2) as "Total_Value_Sold",
        ds.unique_skus as "Number_of_Unique_SKUs_Sold",
        ds.unique_customers as "Count_of_Unique_Customers_Served",
        ds.total_visits as "Number_of_Visits_Made",
        ROUND(
            COALESCE((cd.customers_with_sales * 100.0) / NULLIF(cd.total_customer_contacts, 0), 0), 
            2
        ) as "Conversion_Percentage",
        ROUND(ds.total_minutes, 2) as "Total_Time_Spent_Minutes"
    FROM daily_stats ds
    LEFT JOIN conversion_data cd ON ds.report_date = cd.conv_date
    ORDER BY ds.report_date
    """
    try:
        summary_df = conn.execute(summary_sql).fetchdf()
        create_summary_sheet(wb, summary_df)
        print(f"    Added {len(summary_df)} daily summaries")
    except Exception as e:
        print(f"    Summary sheet error: {e}")
        ws = wb.create_sheet("Summary_Report")
        ws['A1'] = "No summary data available"
    format_excel_file(wb)
    wb.save(f"output/{filename}")
    print(f"    Saved: {filename}")
    return filename

def create_transactions_sheet(wb, df):
    ws = wb.create_sheet("Day_wise_Transactions")
    headers = ["Date", "SKU_Index", "SKU_Sold", "Price_per_SKU", "Quantity_Sold", "Value_Sold"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row_data in df.iterrows():
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            if col_idx == 4:
                cell.number_format = '$#,##0.00'
            elif col_idx == 5:
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.number_format = '$#,##0.00'

def create_summary_sheet(wb, df):
    ws = wb.create_sheet("Summary_Report")
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=2):
        for col_idx, cell in enumerate(row):
            if col_idx == 1:
                cell.number_format = '#,##0'
            elif col_idx == 2:
                cell.number_format = '$#,##0.00'
            elif col_idx in [3, 4, 5]:
                cell.number_format = '#,##0'
            elif col_idx == 6:
                cell.number_format = '0.00"%"'
            elif col_idx == 7:
                cell.number_format = '#,##0.00'

def format_excel_file(wb):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 4, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

def main():
    print("Starting Sales Representative Excel Report Generation")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    if not os.path.exists('resources/assignmentData.csv'):
        print("Error: 'resources/assignmentData.csv' not found!")
        print("   Please ensure the CSV file is in the current directory.")
        return
    files = create_sales_rep_reports()
    print("I am here")
    if files:
        print("\n" + "=" * 60)
        print("REPORT GENERATION COMPLETE!")
        print(f"Total files created: {len(files)}")
        print("\nGenerated Files:")
        for idx, file in enumerate(files, 1):
            if os.path.exists(file):
                file_size = os.path.getsize(file) / 1024
                print(f"  {idx}. {file} ({file_size:.1f} KB)")
            else:
                print(f"  {idx}. {file} (File not found)")
        print("\nAll Excel reports are ready for analysis!")
    else:
        print("\nNo files were created. Please check the error messages above.")

if __name__ == "__main__":
    main()
